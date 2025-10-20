import os
import json
import pandas as pd
from collections import Counter
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def read_phase_results(csv_file):
    
    if os.path.exists(csv_file):
        return pd.read_csv(csv_file)
    return None


def consolidate_results(job_dir):
    
    results_file = os.path.join(job_dir, 'results.json')
    
    if not os.path.exists(results_file):
        return None
    
    with open(results_file, 'r') as f:
        results = json.load(f)
    
    # Read all phase results
    consolidated = {
        'job_id': results['job_id'],
        'timestamp': results['timestamp'],
        'phases': {},
        'statistics': {}
    }
    
    for phase_key, csv_path in results.get('phases', {}).items():
        # Make sure the path is absolute
        if not os.path.isabs(csv_path):
            csv_path = os.path.join(job_dir, csv_path)
        
        df = read_phase_results(csv_path)
        if df is not None:
            consolidated['phases'][phase_key] = df.to_dict('records')
            
            # Calculate statistics for this phase
            if 'prediction' in df.columns or 'Prediction' in df.columns:
                pred_col = 'prediction' if 'prediction' in df.columns else 'Prediction'
                counts = df[pred_col].value_counts().to_dict()
                consolidated['statistics'][phase_key] = counts
    
    return consolidated


def generate_excel_report(job_dir, output_file):

    results_file = os.path.join(job_dir, 'results.json')
    
    if not os.path.exists(results_file):
        raise Exception("Results file not found")
    
    with open(results_file, 'r') as f:
        results = json.load(f)
    
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write each phase to a separate sheet
        for phase_key, csv_path in results.get('phases', {}).items():
            # Make sure the path is absolute
            if not os.path.isabs(csv_path):
                csv_path = os.path.join(job_dir, csv_path)
            
            df = read_phase_results(csv_path)
            if df is not None:
                sheet_name = phase_key.replace('_', ' ').title()
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Format the sheet
                worksheet = writer.sheets[sheet_name]
                
                # Style header row
                header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width


def generate_pdf_report(job_dir, output_file):

    results_file = os.path.join(job_dir, 'results.json')
    
    if not os.path.exists(results_file):
        raise Exception("Results file not found")
    
    with open(results_file, 'r') as f:
        results = json.load(f)
    
    # Create PDF
    pdf = FPDF()
    pdf.add_page()
    
    # Title
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'DeepCovVar Prediction Results', 0, 1, 'C')
    pdf.ln(5)
    
    # Job information
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 8, f"Job ID: {results['job_id']}", 0, 1)
    pdf.cell(0, 8, f"Timestamp: {results['timestamp']}", 0, 1)
    pdf.ln(5)
    
    # Results for each phase
    for phase_key, csv_path in results.get('phases', {}).items():
        # Make sure the path is absolute
        if not os.path.isabs(csv_path):
            csv_path = os.path.join(job_dir, csv_path)
        
        df = read_phase_results(csv_path)
        if df is not None:
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(0, 10, phase_key.replace('_', ' ').title(), 0, 1)
            pdf.ln(2)
            
            # Summary statistics
            pdf.set_font('Arial', '', 10)
            pdf.cell(0, 6, f"Total sequences: {len(df)}", 0, 1)
            
            # If there's a prediction column, show distribution
            if 'prediction' in df.columns or 'Prediction' in df.columns:
                pred_col = 'prediction' if 'prediction' in df.columns else 'Prediction'
                counts = df[pred_col].value_counts()
                pdf.cell(0, 6, "Prediction distribution:", 0, 1)
                for pred, count in counts.items():
                    pdf.cell(0, 6, f"  {pred}: {count} ({count/len(df)*100:.1f}%)", 0, 1)
            
            pdf.ln(5)
    
    # Save PDF
    pdf.output(output_file)


def calculate_statistics(job_dir):

    consolidated = consolidate_results(job_dir)
    
    if not consolidated:
        return None
    
    stats = {
        'total_sequences': 0,
        'phases_completed': len(consolidated['phases']),
        'phase_statistics': consolidated.get('statistics', {})
    }
    
    # Count total sequences from first phase
    if consolidated['phases']:
        first_phase = list(consolidated['phases'].values())[0]
        stats['total_sequences'] = len(first_phase)
    
    return stats
